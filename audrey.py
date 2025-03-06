import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook

st.title("Excel Processing App (Preserve Formulas in Feuil1 & Feuil2)")

# File uploaders for received file & template
uploaded_file = st.file_uploader("📂 Upload the received Excel file", type=["xlsx"])
template_file = st.file_uploader("📂 Upload the template Excel file", type=["xlsx"])

if uploaded_file and template_file:
    # ✅ Load uploaded file into Pandas DataFrame (read first sheet without headers)
    df_uploaded = pd.read_excel(uploaded_file, sheet_name=0, header=None)

    # ✅ Load template file with openpyxl
    wb = load_workbook(template_file)

    # ✅ Step 1: Clear old data (but keep formulas) in **Feuil1 & Feuil2**
    for sheet_name in ["Feuil1", "Feuil2"]:  # Process both sheets
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if not cell.data_type == "f":  # Only clear non-formula cells
                        cell.value = None

    # ✅ Step 2: Copy new uploaded data into **Feuil1** (starting from row 2)
    ws_feuil1 = wb["Feuil1"]
    for i, row in enumerate(df_uploaded.values, start=2):  # Start from row 2 to keep headers
        for j, value in enumerate(row, start=1):
            ws_feuil1.cell(row=i, column=j, value=value)

    # ✅ Step 3: Save updated workbook to a BytesIO buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # ✅ Step 4: Provide a Download Button
    st.download_button(
        label="📥 Download Processed File with Formulas",
        data=output,
        file_name="Processed_Output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
